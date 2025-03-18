#
# В этом модуле находятся функции и константы, которые используются многими другими модулями.
# Здесь не создаётся роутера и нет хэндлеров.
# 

import json
import os.path
import aiogram as ag
import openpyxl
from types import SimpleNamespace

# В "staffToMessage.json" хранятся id аккаунтов сотрудников, соотнесенные с их должностями
with open('staffToMessage.json', encoding='utf-8') as staffToMessageJsonFile:
    staffToMessageData = staffToMessageJsonFile.read()
    # Словарь с id всех нужных для работы бота сотрудников (в основном, используется для отправки им сообщений от имени бота)
    STAFF_TO_MESSAGE = json.loads(staffToMessageData, object_hook=lambda d: SimpleNamespace(**d))

# (1) Для справочника по сотрудникам компании:
# в "staffInfo.json" хранятся данные для справочника по сотрудникам, работающих в компании,
# а также id телеграма для определения того, может ли написавший этому боту пользоваться им
with open('./staffInfo/staffInfo.json', encoding='utf-8') as staffInfoJsonFile:
    staffInfoData = staffInfoJsonFile.read()
    # Сложный объект, состоящий из вложенных словарей, массивов, хранящий информацию о сотрудниках
    STAFF_INFO = json.loads(staffInfoData)
# (2) Для справочника по сотрудникам компании:
# в "departments.json" хранится массив с перечислением отделов компании
with open('./staffInfo/departments.json', encoding='utf-8') as departmentsJsonFile:
    departmentsData = departmentsJsonFile.read()
    # Массив с перечислением отделов компании
    DEPARTMENTS_IN_STAFF_INFO = json.loads(departmentsData)
# (3) Для справочника по сотрудникам компании:
# в "fieldsInStaffInfo" хранится массив с ключами объекта каждого сотрудника в STAFF_INFO 
# (В нём перечисляются только те ключи, информация по которым выводится в справочнике. Служебные ключи не перечислены) 
with open('./staffInfo/keysForEmployee.json', encoding='utf-8') as keysForEmployeeJsonFile:
    keysForEmployeeData = keysForEmployeeJsonFile.read()
    # Массив из ключей словаря для каждого сотрудника в STAFF_INFO
    KEYS_FOR_EMPLOYEE_IN_STAFF_INFO = json.loads(keysForEmployeeData)

# Стандартный "input_field_placeholder" в клавиатурах
STD_INPUT_FIELD_PLACEHOLDER = "Выберите нужный пункт меню:"
# Если пользователь вместо выбора кнопки в клавиатуре отправил что-то некорректное
ANSWER_TO_INCORRECT_CHOICE_IN_KB = "К сожалению, я не могу обработать это сообщение. Вы должны выбрать один из пунктов меню."

# Стандартная клавиатура для ответа на вопрос формата "да/нет"
YES_IN_STD_YES_OR_NO_KB = "✅ Да, всё верно"
NO_IN_STD_YES_OR_NO_KB = "❌ Нет, не хочу"
STD_YES_OR_NO_KB = ag.types.ReplyKeyboardMarkup(
    keyboard=[
        [
            ag.types.KeyboardButton(text=YES_IN_STD_YES_OR_NO_KB),
            ag.types.KeyboardButton(text=NO_IN_STD_YES_OR_NO_KB)
        ],
    ],
    resize_keyboard=True,
    one_time_keyboard=True,
    input_field_placeholder=STD_INPUT_FIELD_PLACEHOLDER
)

# Разбирает сохраненные данные в текущей сессии и возвращает их одной строкой для отправки пользователю.
# ВНИМАНИЕ: строка выделяется курсивом в стиле MARKDOWN, т.е. ставится нижнее подчеркивание в начале и конце;
# чтобы корректно вывести её, необходимо при отправке сообщения указать аргумент parse_mode="MARKDOWN".
def createStringFromStateData(stateData: dict, stateDataKeys: tuple[tuple[str, str]]) -> str:
    messageToSend = "_"
    for i in range(len(stateDataKeys)):
        serviceKey = stateDataKeys[i][0]
        if serviceKey not in stateData: 
            continue
        humanReadableKey = stateDataKeys[i][1]
        messageToSend += humanReadableKey + ":\n" + stateData[serviceKey] + "\n\n"
    messageToSend += "_"
    return messageToSend

# Вводит сохраненные данные текущей сессии в последнюю строку указанного эксель-файла.
# Берёт заголовки полей из stateDataKeys[i][1], данные полей из stateData[stateDataKeys[i][0]].
def enterStateDataToExcel(stateData: dict, stateDataKeys: tuple[tuple[str, str]], excelFilePath: str):
    # Проверяем, существует ли файл
    if os.path.exists(excelFilePath):
        wb = openpyxl.load_workbook(filename=excelFilePath) # Открываем старый файл
    else:
        wb = openpyxl.Workbook() # Создаем новый файл
    ws = wb.active
    # Если лист пустой, добавляем в него заголовки
    if ws.max_row == 1 and ws.max_column == 1:
        # Извлекаем только заголовки (без данных) и складываем их в отдельный массив
        titles = []
        for i in range(len(stateDataKeys)):
            humanReadableKey = stateDataKeys[i][1]
            titles.append(humanReadableKey)
        # Прикрепляем заголовки в последнюю незаполненную строку в файле
        ws.append(titles)
    # Извлекаем только данные (без заголовков) и складываем их в отдельный массив
    data = []
    for i in range(len(stateDataKeys)):
        serviceKey = stateDataKeys[i][0]
        data.append(stateData[serviceKey])
    # Прикрепляем данные в последнюю незаполненную строку в файле
    ws.append(data)
    # Сохраняем и закрываем файл
    wb.save(excelFilePath)
    wb.close()

# Вызывается при успешном прохождении любой из стадий до самого конца
async def executeCorrectEndingOfSession(message: ag.types.Message, state: ag.fsm.context.FSMContext):
    # Завершаем сессию и удаляем данные
    await state.clear()
    # Предлагаем пользователю снова перейти в главное меню
    await message.answer("Чтобы снова вызвать главное меню, воспользуйтесь командой /main_menu.")

# Проверяет, входит ли сотрудник в компанию (находится ли id его телеграма в STAFF_INFO)
def isUserWorkingInCompany(userId) -> bool:
    for dep in DEPARTMENTS_IN_STAFF_INFO:
        for employee in STAFF_INFO[dep]:
            if employee["tg_id"] == userId:
                return True
    return False